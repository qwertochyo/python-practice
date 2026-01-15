import openpyxl
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote

LINKS_PATH = '/Users/andrey/Desktop/code/Python/lab_4/Links.xlsx'
RESULT_PATH = '/Users/andrey/Desktop/code/Python/lab_4/result.xlsx'

def read_links_from_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    links = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            links.append(row[0])
    return links

def extract_imo_mmsi(vsoup):
    imo = mmsi = ''
    tables = vsoup.select('table.details, table.aparams')
    for table in tables:
        for row in table.find_all('tr'):
            cells = [td.text.strip() for td in row.find_all('td')]
            if len(cells) == 2:
                label = cells[0]
                val = cells[1].replace(' ', '')
                if ('IMO' in label or 'MMSI' in label) and '/' in val:
                    parts = val.split('/')
                    if len(parts) == 2:
                        imo, mmsi = parts
                elif 'IMO' in label and not imo:
                    imo = val
                elif 'MMSI' in label and not mmsi:
                    mmsi = val
    return imo, mmsi

def get_vessel_data(link):
    if 'name=' in link:
        base, name = link.split('name=', 1)
        encoded_name = quote(name)
        link = base + 'name=' + encoded_name
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }
    response = requests.get(link, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')
    rows = soup.select('table.results tbody tr')
    if len(rows) != 1:
        return None
    ship_link = rows[0].select_one('a.ship-link')
    vessel_url = 'https://www.vesselfinder.com' + ship_link.get('href') if ship_link else ''
    name = ship_link.select_one('.slna').get_text(strip=True) if ship_link else ''
    vessel_type = ship_link.select_one('.slty').get_text(strip=True) if ship_link else ''
    imo = mmsi = ''
    if vessel_url:
        vresponse = requests.get(vessel_url, headers=headers)
        vsoup = BeautifulSoup(vresponse.text, 'html.parser')
        imo, mmsi = extract_imo_mmsi(vsoup)
    return [name, imo, mmsi, vessel_type]

def write_results_to_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Название', 'IMO', 'MMSI', 'Тип'])
    for row in results:
        ws.append(row)
    wb.save(path)

def main():
    links = read_links_from_excel(LINKS_PATH)
    results = []
    for link in links:
        try:
            data = get_vessel_data(link)
            if data and any(data):
                print(f'[Insert information] Добавлено судно: {data}')
                results.append(data)
        except Exception as e:
            print(f'[ERROR] {link}: {e}')
            pass
    write_results_to_excel(results, RESULT_PATH)
    print(f'Готово! Найдено {len(results)} судов. Результат в {RESULT_PATH}')

if __name__ == '__main__':
    main()